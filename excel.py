from openpyxl import Workbook, load_workbook    #Workbook為EXCEL，openpyxl只支援2010後的EXCEL版本

# wb = load_workbook("excel.xlsx")    #讀取excel 檔案
# ws = wb.active                      #選擇預設的工作表
# print(ws['A5'].value)               #印出A5的資料

# ws['A5'].value = '灰'               #更改A5的內容
# print(ws['A5'].value) 

# wb.save('excel.xlsx')               #儲存檔案，必須把檔案關閉


# print(wb.sheetnames)                #回傳所有工作表的名字

wb = load_workbook("excel.xlsx")
ws = wb['工作表2']

wb.create_sheet('QQ')               #創建工作表
print(wb.sheetnames)                #回傳所有工作表的名字

wb.save('excel.xlsx')