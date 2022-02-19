from openpyxl import Workbook, load_workbook

wb = load_workbook('excel3.xlsx')
ws = wb.active

# ws.merge_cells('A1:E1') #合併儲存格
# ws.unmerge_cells('A1:E1')   #解除合併儲存格
        
# ws.insert_rows(3)   #插入橫排
# ws.insert_cols(4)   #插入直排

ws.delete_rows(3)   #刪除橫排
ws.delete_cols(4)   #刪除直排


wb.save('excel3.xlsx')