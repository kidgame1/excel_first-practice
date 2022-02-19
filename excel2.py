from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = Workbook() #新增一個excel
wb = load_workbook('new_excel.xlsx')
ws = wb.active
# ws.title = 'qq' #預設工作表的名稱

ws.append([123,456,789,0])  #新增一橫排資料
ws.append([123,456,789,0,222])
ws.append([123,456,789,0,688])
ws.append([123,456,789,0,987])

for row in range(1,5):  #讀去&修改範圍資料
    for col in range(1,6):
        char = get_column_letter(col)
        ws[char + str(row)].value = char + str(row)
        


wb.save('new_excel.xlsx')