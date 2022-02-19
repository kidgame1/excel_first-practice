from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = [
    {
        'name': '白',
        'tall': 180,
        'age' : 23,
        'weight': 74
    },
    {
        'name': '黃',
        'tall': 177,
        'age' : 28,
        'weight': 90
    },
    {
        'name': '綠',
        'tall': 160,
        'age' : 30,
        'weight': 60
    },
    {
        'name': '灰',
        'tall': 155,
        'age' : 50,
        'weight': 50
    },
    {
        'name': '黑',
        'tall': 170,
        'age' : 46,
        'weight': 99
    },
]


wb = Workbook()
ws = wb.active

title = ['姓名', '身高', '年紀', '體重']
ws.append(title)

for person in data:
    ws.append(list(person.values()))

for col in range(2,5):
    char = get_column_letter(col)
    ws[char + '7'] = f'=AVERAGE({char + "2"}:{char + "6"})' #平均值

for col in range(1, 5):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True, color="000000FF")


wb.save('data.xlsx')